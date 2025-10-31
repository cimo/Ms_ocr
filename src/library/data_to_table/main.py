import os
import numpy
import pandas
import unicodedata
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

class DataToTable:
    def _isCellInMergedRange(self, cellRow, cellColumn, cellMergeList):
        isMerge = False
        isMergeHorizontal = False

        for rowMin, rowMax, columnMin, columnMax in cellMergeList:
            if rowMin <= cellRow <= rowMax and columnMin <= cellColumn <= columnMax:
                isMerge = True

                if columnMin != columnMax:
                    isMergeHorizontal = True

        return isMerge, isMergeHorizontal

    def _unicodeTextLength(self, text):
        length = 0

        for char in text:
            if unicodedata.east_asian_width(char) in ["F", "W"]:
                length += 2
            elif char.isspace():
                length += 0.5
            else:
                length += 1

        return length
    
    def _html(self):
        resultList = ["<html><head><style>"]
        resultList.append("""
            table {
                border-collapse: collapse;
                table-layout: fixed;
            }
                    
            td {
                border: 1px solid #000000;
                text-align: center;
                vertical-align: middle;
                word-wrap: break-word;
                white-space: pre-wrap;
                font-family: Calibri, sans-serif;
                font-size: 11pt;
                padding: 4px;
            }
        """)
        resultList.append("</style></head><body><table>")

        cellMergeList = []

        for merge in self.mergeList:
            rowStart = merge["row_start"]
            columnStart = merge["column_start"]
            rowEnd = merge["row_end"]
            columnEnd = merge["column_end"]

            cellMergeList.append((rowStart, columnStart, rowEnd, columnEnd))

        cellSkipList = []

        for indexRow, row in enumerate(self.dataFrame.values):
            resultList.append("<tr>")

            for indexColumn, cellValue in enumerate(row):
                if (indexRow, indexColumn) in cellSkipList:
                    continue

                rowSpan, columnSpan = 1, 1
                
                for rowStart, colStart, rowEnd, colEnd in cellMergeList:
                    if indexRow == rowStart and indexColumn == colStart:
                        rowSpan = rowEnd - indexRow + 1
                        columnSpan = colEnd - indexColumn + 1

                        for a in range(rowStart, rowEnd + 1):
                            for b in range(colStart, colEnd + 1):
                                if (a, b) != (indexRow, indexColumn):
                                    cellSkipList.append((a, b))

                        break

                text = str(cellValue).replace("\n", "<br>")

                resultList.append(f'<td rowspan="{rowSpan}" colspan="{columnSpan}">{text}</td>')

            resultList.append("</tr>")

        resultList.append("</table></body></html>")

        with open(self.pathOutput, "w", encoding="utf-8") as file:
            file.write("\n".join(resultList))
    
    def _excel(self):
        writer = pandas.ExcelWriter(self.pathOutput, engine="openpyxl")

        sheetName = "Sheet_1"
            
        self.dataFrame.to_excel(writer, index=False, header=False, sheet_name=sheetName)
        
        worksheet = writer.sheets[sheetName]

        for merge in self.mergeList:
            rowStart = merge["row_start"] + 1
            columnStart = merge["column_start"] + 1
            rowEnd = merge["row_end"] + 1
            columnEnd = merge["column_end"] + 1
            
            if rowStart != rowEnd or columnStart != columnEnd:
                worksheet.merge_cells(start_row=rowStart, start_column=columnStart, end_row=rowEnd, end_column=columnEnd)
        
        cellMergeList = [(cellMergeRange.min_row, cellMergeRange.max_row, cellMergeRange.min_col, cellMergeRange.max_col) for cellMergeRange in worksheet.merged_cells.ranges]

        for indexRow, rowList in enumerate(worksheet.iter_rows(), 1):
            lineTotal = 1

            for cell in rowList:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

                isMerge, _ = self._isCellInMergedRange(cell.row, cell.column, cellMergeList)

                if not isMerge:
                    if isinstance(cell.value, str) and "\n" in cell.value:
                        lineCount = cell.value.count("\n") + 1

                        if lineCount > lineTotal:
                            lineTotal = lineCount

            worksheet.row_dimensions[indexRow].height = max(self.cellHeightDefault * lineTotal, self.cellHeightDefault)
        
        for rowMin, rowMax, columnMin, _ in cellMergeList:
            cell = worksheet.cell(rowMin, columnMin)

            if isinstance(cell.value, str):
                lineCount = cell.value.count("\n") + 1

                heightTotal = self.cellHeightDefault * lineCount
                rowTotal = rowMax - rowMin + 1
                height = heightTotal / rowTotal

                for row in range(rowMin, rowMax + 1):
                    worksheet.row_dimensions[row].height = max(height, self.cellHeightDefault)

        for indexColumn, cellList in enumerate(worksheet.columns, 1):
            textWidthMax = 0

            for cell in cellList:
                _, isMergeHorizontal = self._isCellInMergedRange(cell.row, cell.column, cellMergeList)

                if isMergeHorizontal:
                    continue

                if isinstance(cell.value, str) and len(cell.value) > 0:
                    text = max(cell.value.split("\n"), key=self._unicodeTextLength)
                    textBbox = self.pilFont.getbbox(text)
                    textWidth = textBbox[2] - textBbox[0]
                    
                    if textWidth > textWidthMax:
                        textWidthMax = textWidth

                    cell.font = Font(name="Calibri", size=11)

            if textWidthMax > 0:
                columnWidth =  round(textWidthMax / 7, 2) + 2
                columnLetter = get_column_letter(indexColumn)
                worksheet.column_dimensions[columnLetter].width = columnWidth

        writer.close()

    def _buildDataFrame(self, cellList, rowIndexList, columnIndexList):
        countRow = max(rowIndexList) + 1
        countColumn = max(columnIndexList) + 1

        matrix = [["" for _ in range(countColumn)] for _ in range(countRow)]
        
        cellMergeList = []

        for merge in self.mergeList:
            for a in range(merge["row_start"], merge["row_end"] + 1):
                for b in range(merge["column_start"], merge["column_end"] + 1):
                    if a != merge["row_start"] or b != merge["column_start"]:
                        cellMergeList.append((a, b))

        for index, cell in enumerate(cellList):
            indexRow = rowIndexList[index]
            indexColumn = columnIndexList[index]

            if (indexRow, indexColumn) in cellMergeList:
                continue

            if matrix[indexRow][indexColumn]:
                matrix[indexRow][indexColumn] += " " + cell["text"]
            else:
                matrix[indexRow][indexColumn] = cell["text"]

        result = pandas.DataFrame(matrix)
        result = result.replace("", pandas.NA)
        result = result.dropna(how="all")
        result = result.fillna("")
        result = result.loc[:, (result != "").any(axis=0)]
        result = result.reset_index(drop=True)

        return result

    def _detectCellMerge(self, cellList, rowIndexList, columnIndexList, rowPositionList, columnPositionList):
        resultList = []
        
        for index, cell in enumerate(cellList):
            rowStart = rowIndexList[index]
            columnStart = columnIndexList[index]

            rowEnd = rowStart

            for a in range(rowStart + 1, len(rowPositionList)):
                if cell["y2"] > rowPositionList[a] + self.toleranceY:
                    rowEnd = a
                else:
                    break

            columnEnd = columnStart

            for a in range(columnStart + 1, len(columnPositionList)):
                if cell["x2"] > columnPositionList[a] + self.toleranceX:
                    columnEnd = a
                else:
                    break

            if rowEnd > rowStart or columnEnd > columnStart:
                resultList.append({
                    "row_start": rowStart,
                    "column_start": columnStart,
                    "row_end": rowEnd,
                    "column_end": columnEnd,
                    "text": cell["text"]
                })
        
        return resultList
    
    def _clusterPositionAverage(self, positionList, indexList, count):
        resultList = []

        for a in range(count):
            clusterValue = [positionList[b] for b in range(len(positionList)) if indexList[b] == a]

            resultList.append(numpy.mean(clusterValue) if clusterValue else 0)
        
        return resultList

    def _clusterPosition(self, positionList, tolerance):
        resultList = [0] * len(positionList)
        
        indexSortList = numpy.argsort(positionList)
        valueSortList = numpy.array(positionList)[indexSortList]
        
        resultList[indexSortList[0]] = 0
        
        count = 0
        
        for a in range(1, len(valueSortList)):
            if valueSortList[a] - valueSortList[a - 1] > tolerance:
                count += 1

            resultList[indexSortList[a]] = count

        return resultList

    def _detectGrid(self, cellList):
        topList = [cell["y1"] for cell in cellList]
        leftList = [cell["x1"] for cell in cellList]

        resultRowIndexList = self._clusterPosition(topList, self.toleranceY)
        resultColumnIndexList = self._clusterPosition(leftList, self.toleranceX)

        countRow = max(resultRowIndexList) + 1
        countColumn = max(resultColumnIndexList) + 1

        resultRowPositionList = self._clusterPositionAverage(topList, resultRowIndexList, countRow)
        resultColumnPositionList = self._clusterPositionAverage(leftList, resultColumnIndexList, countColumn)

        return resultRowIndexList, resultColumnIndexList, resultRowPositionList, resultColumnPositionList

    def __init__(self, dataList, pathOutput, pilFont):
        self.pathOutput = pathOutput
        self.pilFont = pilFont
        self.mergeList = []
        self.dataFrame = None
        self.toleranceX = 15
        self.toleranceY = 10
        self.cellHeightDefault = 18

        _, fileExtensionSplit = os.path.splitext(os.path.basename(self.pathOutput))

        cellList = []

        for data in dataList:
            bboxList = data.get("bbox_list", [])
            textList = data.get("text_list", [])

            x1, y1, x2, y2 = bboxList[:4]
            text = "\n".join([text.strip() for text in reversed(textList) if text.strip()])

            cellList.append({
                "x1": x1,
                "y1": y1,
                "x2": x2,
                "y2": y2,
                "width": x2 - x1,
                "height": y2 - y1,
                "text": text.strip()
            })
            cellList.sort(key=lambda cell: (cell["y1"], cell["x1"]))

        rowIndexList, columnIndexList, rowPositionList, columnPositionList = self._detectGrid(cellList)
        self.mergeList = self._detectCellMerge(cellList, rowIndexList, columnIndexList, rowPositionList, columnPositionList)
        self.dataFrame = self._buildDataFrame(cellList, rowIndexList, columnIndexList)

        if fileExtensionSplit.lower() == ".xlsx":
            self._excel()
        elif fileExtensionSplit.lower() == ".html":
            self._html()
