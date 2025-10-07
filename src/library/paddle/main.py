import sys
import os
import shutil
import json
import numpy
import cv2
import pandas
from paddleocr import LayoutDetection, TableClassification, TableCellsDetection, TextDetection, TextRecognition, PaddleOCR
from PIL import Image, ImageDraw, ImageFont
from typing import List, Dict, Tuple
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Source
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from preprocessor import helper as preprocessorHelper

PATH_PADDLE_SYSTEM = "/home/app/.paddlex/"
PATH_PADDLE_FILE_OUTPUT = "/home/app/file/output/paddle/"
PATH_PADDLE_LIBRARY = "/home/app/src/library/paddle/"

pathFont = f"{PATH_PADDLE_SYSTEM}fonts/PingFang-SC-Regular.ttf"

pathModelLayout = f"{PATH_PADDLE_LIBRARY}PP-DocLayout_plus-L/"
pathModelTableClassification = f"{PATH_PADDLE_LIBRARY}PP-LCNet_x1_0_table_cls/"
pathModelTableWired = f"{PATH_PADDLE_LIBRARY}RT-DETR-L_wired_table_cell_det/"
pathModelTableWireless = f"{PATH_PADDLE_LIBRARY}RT-DETR-L_wireless_table_cell_det/"
pathModelTextDetection = f"{PATH_PADDLE_LIBRARY}PP-OCRv5_server_det/"
pathModelTextRecognition = f"{PATH_PADDLE_LIBRARY}PP-OCRv5_server_rec/"

isDebug = True
device = "cpu"

ocr = PaddleOCR(
    text_detection_model_dir=pathModelTextDetection,
    text_detection_model_name="PP-OCRv5_server_det",
    text_recognition_model_dir=pathModelTextRecognition,
    text_recognition_model_name="PP-OCRv5_server_rec",
    use_doc_orientation_classify=False,
    use_doc_unwarping=False,
    use_textline_orientation=False,
    device=device
)

def _inferenceTextDetection(input):
    model = TextDetection(model_dir=pathModelTextDetection, model_name="PP-OCRv5_server_det")
    outputList = model.predict(input=input, batch_size=1)

    for output in outputList:
        if isDebug:
            output.save_to_img(save_path=f"{PATH_PADDLE_FILE_OUTPUT}export/result_detection.jpg")
            output.save_to_json(save_path=f"{PATH_PADDLE_FILE_OUTPUT}export/result_detection.json")

def _inferenceTextRecognition(input):
    model = TextRecognition(model_dir=pathModelTextRecognition, model_name="PP-OCRv5_server_rec")
    outputList = model.predict(input=input, batch_size=1)

    for output in outputList:
        if isDebug:
            output.save_to_img(save_path=f"{PATH_PADDLE_FILE_OUTPUT}export/result_recognition.jpg")
            output.save_to_json(save_path=f"{PATH_PADDLE_FILE_OUTPUT}export/result_recognition.json")

def _filterOverlapBox(boxList):
    resultBoxFilterList = []

    for box in boxList:
        x1, y1, x2, y2 = map(float, box["coordinate"])
        
        isIgnore = False
        
        for boxFilter in resultBoxFilterList:
            xf1, yf1, xf2, yf2 = map(float, boxFilter["coordinate"])

            overlapX = max(0, min(x2, xf2) - max(x1, xf1))
            overlapY = max(0, min(y2, yf2) - max(y1, yf1))

            minWidth = min(x2 - x1, xf2 - xf1)
            minHeight = min(y2 - y1, yf2 - yf1)

            if overlapX / minWidth > 0.9 and overlapY / minHeight > 0.9:
                isIgnore = True

                break

        if isIgnore:
            continue

        resultBoxFilterList.append(box)
    
    return resultBoxFilterList

def _processTable(mode, data, input, count):
    if isDebug:
        inputHeight, inputWidth = input.shape[:2]
        resultImage = Image.new("RGB", (inputWidth, inputHeight), (255, 255, 255))
        imageDraw = ImageDraw.Draw(resultImage)
        resultMergeList = []

    boxList = data.get("boxes", [])

    boxFilterList = _filterOverlapBox(boxList)

    for box in boxFilterList:
        coordinateList = box.get("coordinate", [])
        x1, y1, x2, y2 = map(int, coordinateList)

        imageCrop = input[y1:y2, x1:x2, :]

        resultList = ocr.predict(input=imageCrop)

        for result in resultList:
            coordinateList = [float(a) for a in coordinateList]
            textList = result.get("rec_texts", []) or [""]

            if isDebug:
                lineNumber = max(1, len(textList))
                cellHeight = y2 - y1
                fontSize = max(8, int(cellHeight * 0.6 / lineNumber))
                font = ImageFont.truetype(pathFont, fontSize)

                bboxeList = [imageDraw.textbbox((0, 0), value, font=font) for value in textList]
                textHeightList = [bbox[3] - bbox[1] for bbox in bboxeList]
                textWidthList = [bbox[2] - bbox[0] for bbox in bboxeList]

                totalTextHeight = sum(textHeightList)

                if lineNumber > 1:
                    extraSpace = (cellHeight - totalTextHeight) / (lineNumber + 1)
                else:
                    extraSpace = (cellHeight - totalTextHeight) / 2

                currentY = y1 + extraSpace

                for key, value in enumerate(textList):
                    textWidth = textWidthList[key]
                    textHeight = textHeightList[key]
                    textPositionX = x1 + (x2 - x1 - textWidth) // 2
                    textPositionY = int(currentY)

                    imageDraw.text((textPositionX, textPositionY), value, font=font, fill=(0, 0, 0))

                    currentY += textHeight + extraSpace

                resultMergeList.append({
                    "coordinate": coordinateList,
                    "rec_texts": textList
                })

    if isDebug:
        resultImage.save(f"{PATH_PADDLE_FILE_OUTPUT}table/{mode}/{count}_result.jpg", format="JPEG")

        with open(f"{PATH_PADDLE_FILE_OUTPUT}table/{mode}/{count}_result.json", "w", encoding="utf-8") as file:
            json.dump(resultMergeList, file, ensure_ascii=False, indent=2)

    #DynamicTableExtractor

def _inferenceTableWireless(input, count):
    model = TableCellsDetection(model_dir=pathModelTableWireless, model_name="RT-DETR-L_wireless_table_cell_det", device=device)
    dataList = model.predict(input=input, batch_size=1)

    for data in dataList:
        if isDebug:
            data.save_to_img(save_path=f"{PATH_PADDLE_FILE_OUTPUT}table/wireless/{count}.jpg")
            data.save_to_json(save_path=f"{PATH_PADDLE_FILE_OUTPUT}table/wireless/{count}.json")

        _processTable("wireless", data, input, count)

def _inferenceTableWired(input, count):
    model = TableCellsDetection(model_dir=pathModelTableWired, model_name="RT-DETR-L_wired_table_cell_det", device=device)
    dataList = model.predict(input=input, batch_size=1)

    for data in dataList:
        if isDebug:
            data.save_to_img(save_path=f"{PATH_PADDLE_FILE_OUTPUT}table/wired/{count}.jpg")
            data.save_to_json(save_path=f"{PATH_PADDLE_FILE_OUTPUT}table/wired/{count}.json")

        _processTable("wired", data, input, count)

def _extractTableCell(data, input, count):
    labelNameList = data.get("label_names", [])
    scoreList = data.get("scores", [])

    if len(labelNameList) == len(scoreList):
        resultIndex = int(numpy.argmax(scoreList))
        resultLabel = labelNameList[resultIndex]

        if (resultLabel == "wired_table"):
            if isDebug:
                cv2.imwrite(f"{PATH_PADDLE_FILE_OUTPUT}table/wired/{count}_crop.jpg", input)

            _inferenceTableWired(input, count)
        elif (resultLabel == "wireless_table"):
            if isDebug:
                cv2.imwrite(f"{PATH_PADDLE_FILE_OUTPUT}table/wireless/{count}_crop.jpg", input)

            _inferenceTableWireless(input, count)

def _inferenceTableClassification(input, count):
    model = TableClassification(model_dir=pathModelTableClassification, model_name="PP-LCNet_x1_0_table_cls", device=device)
    dataList = model.predict(input=input, batch_size=1)

    for data in dataList:
        if isDebug:
            data.save_to_img(save_path=f"{PATH_PADDLE_FILE_OUTPUT}table/classification/{count}.jpg")
            data.save_to_json(save_path=f"{PATH_PADDLE_FILE_OUTPUT}table/classification/{count}.json")

    _extractTableCell(data, input, count)

def _extractTable(data, input):
    boxList = data.get("boxes", [])

    count = 0

    for box in boxList:
        if str(box.get("label", "")).lower() != "table":
            continue

        coordinateList = box.get("coordinate", [])
        x1, y1, x2, y2 = map(int, coordinateList)

        inputCrop = input[y1:y2, x1:x2, :]

        _inferenceTableClassification(inputCrop, count)

        count += 1

def removeOutputDir():
    if os.path.exists(f"{PATH_PADDLE_FILE_OUTPUT}layout/"):
        shutil.rmtree(f"{PATH_PADDLE_FILE_OUTPUT}layout/")

    if os.path.exists(f"{PATH_PADDLE_FILE_OUTPUT}table/"):
        shutil.rmtree(f"{PATH_PADDLE_FILE_OUTPUT}table/")

    if os.path.exists(f"{PATH_PADDLE_FILE_OUTPUT}export/"):
        shutil.rmtree(f"{PATH_PADDLE_FILE_OUTPUT}export/")

def createOutputDir():
    os.makedirs(f"{PATH_PADDLE_FILE_OUTPUT}layout/", exist_ok=True)
    os.makedirs(f"{PATH_PADDLE_FILE_OUTPUT}table/classification/", exist_ok=True)
    os.makedirs(f"{PATH_PADDLE_FILE_OUTPUT}table/wired/", exist_ok=True)
    os.makedirs(f"{PATH_PADDLE_FILE_OUTPUT}table/wireless/", exist_ok=True)
    os.makedirs(f"{PATH_PADDLE_FILE_OUTPUT}export/", exist_ok=True)

def inferenceLayout(input):
    model = LayoutDetection(model_dir=pathModelLayout, model_name="PP-DocLayout_plus-L", device=device)
    dataList = model.predict(input=input, batch_size=1)

    for data in dataList:
        if isDebug:
            data.save_to_img(save_path=f"{PATH_PADDLE_FILE_OUTPUT}layout/result.jpg")
            data.save_to_json(save_path=f"{PATH_PADDLE_FILE_OUTPUT}layout/result.json")

        _extractTable(data, input)

class JsonToExcel:
    def _clusterPosition(self, positionList, tolerance):
        indexSortList = numpy.argsort(positionList)
        valueSortList = numpy.array(positionList)[indexSortList]

        resultList = [0] * len(positionList)
        
        resultList[indexSortList[0]] = 0

        count = 0
        
        for a in range(1, len(valueSortList)):
            if valueSortList[a] - valueSortList[a - 1] > tolerance:
                count += 1

            resultList[indexSortList[a]] = count
        
        return resultList

    def _clusterPositionAverage(self, positionList, indexList, count):
        resultList = []

        for a in range(count):
            clusterValues = [positionList[b] for b in range(len(positionList)) if indexList[b] == a]
            
            resultList.append(numpy.mean(clusterValues) if clusterValues else 0)
            
        return resultList

    def _detectGrid(self, cellList):
        centerYlist = [cell["center_y"] for cell in cellList]
        leftXlist = [cell["x1"] for cell in cellList]

        resultRowIndexList = self._clusterPosition(centerYlist, self.toleranceY)
        resultColumnIndexList = self._clusterPosition(leftXlist, self.toleranceX)

        countRow = max(resultRowIndexList) + 1
        countColumn = max(resultColumnIndexList) + 1
        
        resultRowPositionList = self._clusterPositionAverage(centerYlist, resultRowIndexList, countRow)
        resultColumnPositionList = self._clusterPositionAverage(leftXlist, resultColumnIndexList, countColumn)
        
        return resultRowIndexList, resultColumnIndexList, resultRowPositionList, resultColumnPositionList
    
    def _detectMergedCell(self, cellList, rowIndexList, columnIndexList, columnPositionList):
        resultList = []
        
        if len(columnPositionList) < 2:
            return resultList
        
        for index, cell in enumerate(cellList):
            indexRow = rowIndexList[index]
            indexColumnStart = columnIndexList[index]
            indexColumnEnd = indexColumnStart

            for a in range(indexColumnStart + 1, len(columnPositionList)):
                if columnPositionList[a] < cell["x2"] - self.toleranceX:
                    indexColumnEnd = a
                else:
                    break

            if indexColumnEnd == indexColumnStart:
                continue

            resultList.append({
                "row": indexRow,
                "column_start": indexColumnStart,
                "column_end": indexColumnEnd,
                "text": cell["text"]
            })

        return resultList
    
    def _buildTableMatrix(self, cellList, rowIndexList, columnIndexList, cellMergedList):
        countRow = max(rowIndexList) + 1
        countColumn = max(columnIndexList) + 1

        matrix = [["" for _ in range(countColumn)] for _ in range(countRow)]

        cellMergeDict = {}

        for cellMerged in cellMergedList:
            key = (cellMerged["row"], cellMerged["column_start"])
            cellMergeDict[key] = cellMerged

        for index, cell in enumerate(cellList):
            indexRow = rowIndexList[index]
            indexColumn = columnIndexList[index]
            keyMerge = (indexRow, indexColumn)

            if keyMerge in cellMergeDict:
                cellMergeStatusList = cellMergeDict[keyMerge]

                matrix[indexRow][indexColumn] = cell["text"]
                
                self.cellMergeStatusList.append({
                    "row": indexRow,
                    "column_start": indexColumn,
                    "column_end": cellMergeStatusList["column_end"]
                })
            else:
                if matrix[indexRow][indexColumn]:
                    matrix[indexRow][indexColumn] += " " + cell["text"]
                else:
                    matrix[indexRow][indexColumn] = cell["text"]

        dataFrame = pandas.DataFrame(matrix)

        dataFrame = dataFrame.replace("", pandas.NA)
        dataFrame = dataFrame.dropna(how="all")
        dataFrame = dataFrame.fillna("")
        dataFrame = dataFrame.loc[:, (dataFrame != "").any(axis=0)]
        dataFrame = dataFrame.reset_index(drop=True)
        
        return dataFrame

    def __init__(self, isDebug, dataList, outputPath):
        self.toleranceX = 10
        self.toleranceY = 15
        self.cellMergeStatusList = []

        sheetName = "Sheet1"
        cellList = []

        for data in dataList:
            coorinateList = data.get("coordinate", [])
            textList = data.get("rec_texts", [])

            if len(coorinateList) >= 4:
                x1, y1, x2, y2 = coorinateList[:4]
                text = "\n".join([text.strip() for text in textList if text.strip()])

                cellList.append({
                    "x1": x1,
                    "y1": y1,
                    "x2": x2,
                    "y2": y2,
                    "text": text.strip(),
                    "center_x": (x1 + x2) / 2,
                    "center_y": (y1 + y2) / 2,
                    "width": x2 - x1,
                    "height": y2 - y1
                })

        rowIndexList, columnIndexList, rowPositionList, columnPositionList = self._detectGrid(cellList)

        cellMergedList = self._detectMergedCell(cellList, rowIndexList, columnIndexList, columnPositionList)
        
        dataFrame = self._buildTableMatrix(cellList, rowIndexList, columnIndexList, cellMergedList)

        with pandas.ExcelWriter(outputPath, engine="openpyxl") as writer:
            dataFrame.to_excel(writer, index=False, header=False, sheet_name=sheetName)
            
            worksheet = writer.sheets[sheetName]

            for rowList in worksheet.iter_rows():
                for cell in rowList:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cellmergeStatus in self.cellMergeStatusList:
                row = cellmergeStatus["row"] + 1
                columnStart = cellmergeStatus["column_start"] + 1
                columnEnd = cellmergeStatus["column_end"] + 1

                cellStart = f"{get_column_letter(columnStart)}{row}"
                cellEnd = f"{get_column_letter(columnEnd)}{row}"

                worksheet.merge_cells(f"{cellStart}:{cellEnd}")
                worksheet[cellStart].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for index, columnList in enumerate(worksheet.columns, 1):
                maxLength = 0

                columnLetter = worksheet.cell(row=1, column=index).column_letter
                
                for cell in columnList:
                    cellLength = len(str(cell.value))

                    if cellLength > maxLength:
                        maxLength = cellLength

                worksheet.column_dimensions[columnLetter].width = min(maxLength + 3, 50)
        
        if isDebug:
            print(f"✓ Table: {len(dataFrame)} row - {len(dataFrame.columns)} column.")

            if self.cellMergeStatusList:
                print(f"✓ Merged cell: {len(self.cellMergeStatusList)}.")

def Main():
    # Execution
    #removeOutputDir()

    #createOutputDir()

    #image = preprocessorHelper.open("/home/app/file/input/1_jp.jpg")
    #_, _, _, imageResize, _ = preprocessorHelper.resize(image, 2048)

    #inferenceLayout(imageResize)

    with open(f"{PATH_PADDLE_FILE_OUTPUT}table/wireless/0_result.json", "r", encoding="utf-8") as file:
        dataList = json.load(file)

    JsonToExcel(isDebug, dataList, f"{PATH_PADDLE_FILE_OUTPUT}table/wireless/0_result.xlsx")

Main()
