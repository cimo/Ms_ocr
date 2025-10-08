import numpy
import pandas
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class JsonToExcel:
    def _clusterPosition(self, positionList, tolerance):
        resultList = [0] * len(positionList)
        
        indexSortList = numpy.argsort(positionList)
        valueSortList = numpy.array(positionList)[indexSortList]
        
        resultList[indexSortList[0]] = 0
        
        count = 0
        
        for a in range(1, len(valueSortList)):
            if valueSortList[a] - valueSortList[a - 1] > tolerance:
                count += 1

            resultList[indexSortList[a]] = count

        return resultList

    def _clusterPositionAverage(self, positionList, indexList, count):
        resultList = []

        for a in range(count):
            clusterValues = [positionList[b] for b in range(len(positionList)) if indexList[b] == a]

            resultList.append(numpy.mean(clusterValues) if clusterValues else 0)
        
        return resultList

    def _detectGrid(self, cellList):
        centerYlist = [cell["y1"] for cell in cellList]
        leftXlist = [cell["x1"] for cell in cellList]

        resultRowIndexList = self._clusterPosition(centerYlist, self.toleranceY)
        resultColumnIndexList = self._clusterPosition(leftXlist, self.toleranceX)

        countRow = max(resultRowIndexList) + 1
        countColumn = max(resultColumnIndexList) + 1

        resultRowPositionList = self._clusterPositionAverage(centerYlist, resultRowIndexList, countRow)
        resultColumnPositionList = self._clusterPositionAverage(leftXlist, resultColumnIndexList, countColumn)

        return resultRowIndexList, resultColumnIndexList, resultRowPositionList, resultColumnPositionList

    def _detectCellMerge(self, cellList, rowIndexList, columnIndexList, rowPositionList, columnPositionList):
        resultList = []
        
        for index, cell in enumerate(cellList):
            rowStart = rowIndexList[index]
            columnStart = columnIndexList[index]

            rowEnd = rowStart

            for a in range(rowStart + 1, len(rowPositionList)):
                if cell["y2"] > rowPositionList[a] + self.toleranceY:
                    rowEnd = a
                else:
                    break

            columnEnd = columnStart

            for a in range(columnStart + 1, len(columnPositionList)):
                if cell["x2"] > columnPositionList[a] + self.toleranceX:
                    columnEnd = a
                else:
                    break

            if rowEnd > rowStart or columnEnd > columnStart:
                resultList.append({
                    "row_start": rowStart,
                    "column_start": columnStart,
                    "row_end": rowEnd,
                    "column_end": columnEnd,
                    "text": cell["text"]
                })
        
        return resultList
    
    def _buildTableMatrix(self, cellList, rowIndexList, columnIndexList, mergeList):
        countRow = max(rowIndexList) + 1
        countColumn = max(columnIndexList) + 1

        matrix = [["" for _ in range(countColumn)] for _ in range(countRow)]
        
        cellMergeList = []

        for merge in mergeList:
            for a in range(merge["row_start"], merge["row_end"] + 1):
                for b in range(merge["column_start"], merge["column_end"] + 1):
                    if a != merge["row_start"] or b != merge["column_start"]:
                        cellMergeList.append((a, b))

        for index, cell in enumerate(cellList):
            indexRow = rowIndexList[index]
            indexColumn = columnIndexList[index]

            if (indexRow, indexColumn) in cellMergeList:
                continue

            if matrix[indexRow][indexColumn]:
                matrix[indexRow][indexColumn] += " " + cell["text"]
            else:
                matrix[indexRow][indexColumn] = cell["text"]

        dataFrame = pandas.DataFrame(matrix)
        dataFrame = dataFrame.replace("", pandas.NA)
        dataFrame = dataFrame.dropna(how="all")
        dataFrame = dataFrame.fillna("")
        dataFrame = dataFrame.loc[:, (dataFrame != "").any(axis=0)]
        dataFrame = dataFrame.reset_index(drop=True)

        return dataFrame

    def __init__(self, isDebug, dataList, outputPath):
        self.toleranceX = 15
        self.toleranceY = 10

        sheetName = "Sheet1"
        cellList = []

        for data in dataList:
            bboxList = data.get("bbox_list", [])
            textList = data.get("text_list", [])

            if len(bboxList) >= 4:
                x1, y1, x2, y2 = bboxList[:4]
                text = "\n".join([text.strip() for text in textList if text.strip()])

                cellList.append({
                    "x1": x1,
                    "y1": y1,
                    "x2": x2,
                    "y2": y2,
                    "width": x2 - x1,
                    "height": y2 - y1,
                    "text": text.strip()
                })

        rowIndexList, columnIndexList, rowPositionList, columnPositionList = self._detectGrid(cellList)
        mergeList = self._detectCellMerge(cellList, rowIndexList, columnIndexList, rowPositionList, columnPositionList)
        dataFrame = self._buildTableMatrix(cellList, rowIndexList, columnIndexList, mergeList)

        with pandas.ExcelWriter(outputPath, engine="openpyxl") as writer:
            dataFrame.to_excel(writer, index=False, header=False, sheet_name=sheetName)
            worksheet = writer.sheets[sheetName]

            for merge in mergeList:
                rowStart = merge["row_start"] + 1
                columnStart = merge["column_start"] + 1
                rowEnd = merge["row_end"] + 1
                columnEnd = merge["column_end"] + 1
                
                if rowStart != rowEnd or columnStart != columnEnd:
                    worksheet.merge_cells(start_row=rowStart, start_column=columnStart, end_row=rowEnd, end_column=columnEnd)

            for index, columnList in enumerate(worksheet.columns, 1):
                lengthMax = 0
                columnLetter = get_column_letter(index)

                for cell in columnList:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    if cell.value:
                        lineSplit = str(cell.value).split("\n")

                        for line in lineSplit:
                            lengthFixed = 0
                            
                            for a in line:
                                if ord(a) > 255:
                                    lengthFixed += 2
                                else:
                                    lengthFixed += 1

                            lengthMax = max(lengthMax, lengthFixed)

                worksheet.column_dimensions[columnLetter].width = lengthMax * 1.2

        if isDebug:
            print(f"✓ Table: {len(dataFrame)} row - {len(dataFrame.columns)} column.")
            print(f"✓ Cell merge: {len(mergeList)}.")
